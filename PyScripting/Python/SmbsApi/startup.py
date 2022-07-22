from clr_loader import get_coreclr
from pythonnet import set_runtime
import os
import sys

sys.path.append(r'C:\\Users\musa.mohamma\Documents\R&D\smbs_scripting\smbs-be\SMBS.Server\SMBS.Shared\bin\Debug\netcoreapp3.1')
sys.path.append(r'C:\Users\musa.mohamma\Documents\R&D\smbs_scripting\smbs-be\SMBS.Server\SMBS.Engine\bin\Debug\netcoreapp3.1')
set_runtime(get_coreclr(r'C:\Users\musa.mohamma\Documents\R&D\smbs_scripting\smbs-be\SMBS.Server\SMBS.Shared\bin\Debug\netcoreapp3.1\SMBS.Shared.runtimeconfig.json'))
                        
import clr                                                  
clr.AddReference('SMBS.Shared')                           
clr.AddReference('SMBS.Engine')                  
import System

import clr
clr.AddReference("SMBS.Engine")
clr.AddReference("SMBS.Shared")

import System

from SMBS.Shared.DataImport import SingleInput, PvtData, PvtDataRow, LabPVT, RockData, TankData
from SMBS.Shared.DataImport import AquiferData, RelativePermeabilityData
from SMBS.Shared.DataImport import RelativePermeabilityDataRow
from SMBS.Shared.DataImport import ProductionData, ProductionDataRow
from SMBS.Shared import *
from SMBS.Engine import SetUp


from openpyxl import Workbook, load_workbook



def convert_python_datetime_to_csharp_datetime(python_datetime):
    d = python_datetime
    return System.DateTime(d.year, d.month, d.day, d.hour, d.minute, d.second)

def to_csharp_int_list(lst):
    lstcs = System.Collections.Generic.List[System.Int32]()
    for e in lst:
        lstcs.Add(e)
    return lstcs


# Tank Data
def _read_tank_data(ws):
    convert = convert_python_datetime_to_csharp_datetime
    tnk = TankData()
    tnk.StartDateOfProduction = convert(ws['B2'].value)
    tnk.FluidType = ws['B3'].value
    tnk.InitialPressure = ws['B4'].value
    tnk.ConnateWaterSaturation = ws['B5'].value

    props = ['GasCap', 'StandardOilInitiallyInPlace', 'StandardGasInitiallyInPlace', 
             'Thickness', 'Length', 'Width', 'Radius']

    rows_gen = ws.iter_rows(min_row=7, min_col=2, max_col=3, max_row=13, values_only=True)

    for prop, row in zip(props, rows_gen):
        low, up = row
        setattr(tnk, prop+'LowerBound', low)
        setattr(tnk, prop+'UpperBound', up)
        
    
    # tank production data
    min_row, max_row = 4, 11
    min_col, max_col = 7, 13

    props = ["Time", "Pressure", "OilProduced", "GasProduced", 
             "WaterProduced", "GasInjected", "WaterInjected"]

    rows_gen = ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col, max_row=max_row, 
                            values_only=True)
    prod_data = ProductionData()
    for row in rows_gen:
        row_data = ProductionDataRow()
        for prop, col in zip(props, row):
            if prop == "Time":
                setattr(row_data, prop, convert_python_datetime_to_csharp_datetime(col))
            else:
                setattr(row_data, prop, col)
        prod_data.Add(row_data)

    tnk.ProductionData = prod_data
    
    
    # Tank data Lab PVT Data
    pvt = LabPVT()
    SurfacePVT = PvtDataRow()      #SurfacePVT()
    BubblePointPVT = PvtDataRow()  #BubblePointPVT()
    DewPointPVT = PvtDataRow()     #DewPointPVT()

    objs = [SurfacePVT, BubblePointPVT, DewPointPVT]
    col_gen = ws.iter_cols(min_row=39, min_col=2, max_col=4, max_row=43, values_only=True)
    for obj, col in zip(objs, col_gen):
        obj.Pressure, obj.Temperature, obj.GasOilRatio, obj.OilDensity, obj.GasDensity = col

    pvt.SurfacePVT = SurfacePVT
    pvt.BubblePointPVT = BubblePointPVT
    pvt.DewPointPVT = DewPointPVT

    tnk.LabPVTMeasurement = pvt    
    
    return tnk


# Rock Data
def _read_rock_data(ws):
    rows_gen = ws.iter_rows(min_row=28, min_col=2, max_col=3, max_row=29, values_only=True)
    rows = tuple(rows_gen)
    rock_data = RockData()
    rock_data.PermeabilityLowerBound, rock_data.PermeabilityUpperBound = rows[0]
    rock_data.PorosityLowerBound, rock_data.PorosityUpperBound = rows[1]
    return rock_data


# Aquifer Data
def _read_aquifer_data(ws):
    aq = AquiferData()
    aq.Position = ws.cell(16, 2).value
    aq.Geometry = ws.cell(17, 2).value
    aq.ExternalBoundaryCondition = ws.cell(18, 2).value
    aq.WaterInfluxModel = ws.cell(19, 2).value

    props = ["OuterInnerRadius", "EncroachmentAngle", "Volume", "Anisotropy"]

    rows_gen = ws.iter_rows(min_row=21, min_col=2, max_col=3, max_row=24, values_only=True)
    for prop, row in zip(props, rows_gen):
        low, up = row
        setattr(aq, prop+'LowerBound', low)
        setattr(aq, prop+'UpperBound', up)
        
    return aq


# RelativePermeabilityData
def _read_relperm_data(ws):
    relperm = RelativePermeabilityData()
    props = ["OilRelPerm", "WaterRelPerm", "GasRelPerm"]

    rows_gen = ws.iter_rows(min_row=33, min_col=2, max_col=4, max_row=35, values_only=True)

    for prop, row in zip(props, rows_gen):
        obj = RelativePermeabilityDataRow()
        obj.ResidualSaturation, obj.EndPoint, obj.Exponent = row
        setattr(relperm, prop, obj)
    
    return relperm


# PVT Data
def _read_extpvt_data(ws):
    pvt_data = PvtData()
    min_row, max_row = 7, 357
    min_col, max_col = 1, 18

    for row in range(min_row, max_row+1):
        pvt_row = PvtDataRow()
        for col in range(min_col, max_col+1):
            header1 = ws.cell(3, col).value
            header2 = ws.cell(4, col).value or ""
            prop = (header1 + header2).replace(" ", "")
            if prop == 'WaterCompress.': prop = 'WaterCompressibility'
            if not hasattr(pvt_row, prop):
                raise ValueError(f"header {header1 + header2} is invalid")
            setattr(pvt_row, prop, ws.cell(row, col).value)
        pvt_data.Add(pvt_row)
        
    return pvt_data


def create_single_input_from_excel(xlfile: str,
                                   tanksheet: str = "Tank",
                                   pvtsheet: str = "ExtPVT") -> SingleInput:
    wb = load_workbook(xlfile)
    tank_data = _read_tank_data(wb[tanksheet])
    rock_data = _read_rock_data(wb[tanksheet])
    aquifer_data = _read_aquifer_data(wb[tanksheet])
    relperm_data = _read_relperm_data(wb[tanksheet])
    extpvt_data = _read_extpvt_data(wb[pvtsheet])

    inp = SingleInput()
    inp.TankData = tank_data
    inp.RockData = rock_data
    inp.AquiferData = aquifer_data
    inp.RelativePermeabilityData = relperm_data
    inp.PvtData = extpvt_data

    return inp
    
